#!/usr/bin/env python3
"""Generate data/all_deletions.xlsx — one sheet per AC, restricted columns excluded."""
import os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

AC_ROOT = os.path.join(os.path.dirname(__file__), '..', 'AC')
OUT = os.path.join(os.path.dirname(__file__), '..', 'data', 'all_deletions.xlsx')

EXCLUDE = {'religion', 'ocr_pass', 'warning', 'repeat'}
KEEP = ['phase', 'pdf_file', 'part_no', 'page_no', 'box_no', 'epic', 'name', 'gender']

HEADER_FILL = PatternFill('solid', fgColor='1a1a2e')
HEADER_FONT = Font(bold=True, color='B9BF1F', name='Courier New', size=10)
DATA_FONT = Font(name='Courier New', size=9)

def dir_to_sheet(dirname):
    parts = dirname.split('_', 1)
    ac_no = parts[0]
    ac_name = parts[1].replace('_', ' ') if len(parts) > 1 else ''
    return f"{ac_no}-{ac_name}"[:31]

def main():
    ac_dirs = sorted([
        d for d in os.listdir(AC_ROOT)
        if os.path.isdir(os.path.join(AC_ROOT, d)) and not d.startswith('.')
    ], key=lambda x: int(x.split('_')[0]) if x.split('_')[0].isdigit() else 9999)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for dirname in ac_dirs:
        csv_path = os.path.join(AC_ROOT, dirname, 'deletion.csv')
        if not os.path.exists(csv_path):
            print(f"  SKIP {dirname} — no deletion.csv")
            continue

        sheet_name = dir_to_sheet(dirname)
        ws = wb.create_sheet(title=sheet_name)

        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            cols = [c for c in KEEP if c in (reader.fieldnames or [])]

            for col_i, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=col_i, value=col.upper())
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center')

            row_i = 2
            for row in reader:
                for col_i, col in enumerate(cols, 1):
                    ws.cell(row=row_i, column=col_i, value=row.get(col, '')).font = DATA_FONT
                row_i += 1

        ws.freeze_panes = 'A2'
        print(f"  {dirname}: {row_i-2} rows → sheet '{sheet_name}'")

    wb.save(OUT)
    print(f"\nSaved → {OUT}")

if __name__ == '__main__':
    main()
